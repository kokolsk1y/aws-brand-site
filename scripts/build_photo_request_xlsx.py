"""
build_photo_request_xlsx.py — формирует Excel-файл для запроса фото у поставщика.

Структура: 3 листа (УНО / АУРА / ДИЗАЙН) + общий лист «Сводка».
Подсветка групп цветом, фиксированная шапка, ширины колонок, гиперссылки на stv.
"""

from __future__ import annotations
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SERIES_JSON = ROOT / "src" / "data" / "series.json"
STV_JSON = ROOT / "scripts" / "audit" / "stv_data.json"
OUT_XLSX = ROOT / "scripts" / "audit" / "photo_request.xlsx"

SERIES_RU = {"uno": "УНО", "aura": "АУРА", "design": "ДИЗАЙН"}
GROUP_RU = {
    "switches": "Выключатели",
    "sockets": "Розетки",
    "frames": "Рамки",
    "accessories": "Аксессуары",
    "other": "Прочее",
}
GROUP_ORDER = ["switches", "sockets", "frames", "accessories", "other"]
COLOR_RU = {
    "white": "Белый", "black": "Чёрный", "grey": "Серый",
    "dark_grey": "Тёмно-серый", "silver": "Серебро", "gold": "Золото",
    "matte_black": "Чёрный матовый", "mahogany": "Махагон",
    "bronze": "Бронза", "chrome": "Хром", "aluminium": "Алюминий",
}

# Цветовая тема серий (мягкая заливка для заголовков групп)
SERIES_FILL = {
    "УНО": "FFE8F0E1",
    "АУРА": "FFFFF4E0",
    "ДИЗАЙН": "FFE6EEF7",
}
SERIES_ACCENT = {
    "УНО": "FF8CB86B",
    "АУРА": "FFD4A24C",
    "ДИЗАЙН": "FF5B7B9C",
}

HEADERS = ["Серия", "Группа", "Артикул", "Название", "Цвет", "Ссылка на stv39",
           "Фото получено", "Файл / папка", "Примечание"]
WIDTHS = [10, 13, 22, 50, 16, 38, 14, 22, 28]


def thin_border() -> Border:
    s = Side(style="thin", color="FFDDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(cell, fill_color: str = "FF111111"):
    cell.font = Font(name="Inter", size=11, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_group_header(cell, fill_color: str, font_color: str = "FF111111"):
    cell.font = Font(name="Inter", size=11, bold=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_data(cell, is_alt: bool = False):
    cell.font = Font(name="Inter", size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border()
    if is_alt:
        cell.fill = PatternFill("solid", fgColor="FFF9F9F9")


def _covered_articles() -> set:
    """Артикулы, на которые фото уже получены (есть в раскладке поставщика)."""
    try:
        from verify_photos import build_matches
        matched, *_ = build_matches()
        return {article for (_skey, article) in matched}
    except Exception as e:
        print(f"[warn] не удалось получить покрытые артикулы: {e}")
        return set()


def collect_rows() -> list[dict]:
    series = json.loads(SERIES_JSON.read_text(encoding="utf-8"))
    stv = json.loads(STV_JSON.read_text(encoding="utf-8"))
    url_map = {it["article"]: it.get("url", "") for sk in stv for it in stv[sk]["items"]}
    covered = _covered_articles()

    rows = []
    for skey in ("uno", "aura", "design"):
        for gkey in GROUP_ORDER:
            for it in series[skey]["groups"].get(gkey, []):
                # новый товар без фото: price=null И фото ещё НЕ прислали
                if it.get("price") is None and it["article"] not in covered:
                    rows.append({
                        "series": SERIES_RU[skey],
                        "group": GROUP_RU[gkey],
                        "article": it["article"],
                        "name": it["name"],
                        "color": COLOR_RU.get(it.get("color", ""), it.get("color", "")),
                        "url": url_map.get(it["article"], ""),
                    })
    return rows


def fill_sheet(ws, rows: list[dict], title: str) -> None:
    ws.title = title
    ws.freeze_panes = "A2"

    # шапка
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c)
    ws.row_dimensions[1].height = 36

    # ширины
    for col, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    row_idx = 2
    prev_group = None
    prev_series = None
    is_alt = False

    for r in rows:
        # вставляем разделитель группы
        if (r["series"], r["group"]) != (prev_series, prev_group):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(HEADERS))
            cell = ws.cell(row=row_idx, column=1, value=f"  {r['series']} → {r['group']}")
            style_group_header(cell, SERIES_FILL[r["series"]])
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1
            prev_series, prev_group = r["series"], r["group"]
            is_alt = False

        values = [r["series"], r["group"], r["article"], r["name"], r["color"],
                  r["url"], "", "", ""]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            style_data(c, is_alt)
        # ссылка как гиперссылка
        if r["url"]:
            link_cell = ws.cell(row=row_idx, column=6)
            link_cell.hyperlink = r["url"]
            link_cell.font = Font(name="Inter", size=10, color="FF1B5FBE", underline="single")
            link_cell.value = "открыть"
        row_idx += 1
        is_alt = not is_alt


def fill_summary(ws, rows: list[dict]) -> None:
    ws.title = "Сводка"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    info = [
        ("Файл", "Запрос фотографий — новые артикулы AWS"),
        ("Дата", "2026-05-28"),
        ("Источник артикулов", "stv39.ru (Электроцентр)"),
        ("Всего артикулов", len(rows)),
        ("", ""),
        ("Серия УНО", sum(1 for r in rows if r["series"] == "УНО")),
        ("Серия АУРА", sum(1 for r in rows if r["series"] == "АУРА")),
        ("Серия ДИЗАЙН", sum(1 for r in rows if r["series"] == "ДИЗАЙН")),
        ("", ""),
        ("Как пользоваться", ""),
        ("", "1. Откройте лист нужной серии (УНО / АУРА / ДИЗАЙН)."),
        ("", "2. По колонке «Ссылка на stv39» можно сразу перейти к товару."),
        ("", "3. Когда фото получено — отметьте в колонке «Фото получено» (можно «v» или дату)."),
        ("", "4. В колонке «Файл / папка» можно указать имя файла/папки на диске."),
        ("", "5. Примечание — любые комментарии (например, «фото в плохом качестве, пересняли»)."),
    ]
    for i, (k, v) in enumerate(info, 1):
        ka = ws.cell(row=i, column=1, value=k)
        va = ws.cell(row=i, column=2, value=v)
        if k and not isinstance(v, int) and i <= 9:
            ka.font = Font(name="Inter", size=11, bold=True)
            va.font = Font(name="Inter", size=11)
        elif isinstance(v, int):
            ka.font = Font(name="Inter", size=11, bold=True)
            va.font = Font(name="Inter", size=14, bold=True, color="FF111111")
        else:
            ka.font = Font(name="Inter", size=10, italic=True, color="FF666666")
            va.font = Font(name="Inter", size=10, color="FF333333")
        ka.alignment = Alignment(vertical="center")
        va.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 22 if not (k == "Как пользоваться" or (not k and i > 9)) else 20


def main() -> None:
    rows = collect_rows()
    wb = Workbook()

    # лист «Сводка» первым
    fill_summary(wb.active, rows)

    # три листа серий
    for series in ("УНО", "АУРА", "ДИЗАЙН"):
        ws = wb.create_sheet(series)
        s_rows = [r for r in rows if r["series"] == series]
        fill_sheet(ws, s_rows, series)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    # markdown-копия (для удобного просмотра в редакторе)
    out_md = OUT_XLSX.with_suffix(".md")
    lines = ["# Запрос фото — осталось дозапросить у поставщика\n",
             f"Всего артикулов без фото: **{len(rows)}**\n"]
    for s in ("УНО", "АУРА", "ДИЗАЙН"):
        s_rows = [r for r in rows if r["series"] == s]
        lines.append(f"\n## {s} — {len(s_rows)}\n")
        prev_g = None
        for r in s_rows:
            if r["group"] != prev_g:
                lines.append(f"\n### {r['group']}\n")
                lines.append("| Артикул | Название | Цвет |")
                lines.append("|---|---|---|")
                prev_g = r["group"]
            lines.append(f"| {r['article']} | {r['name']} | {r['color']} |")
    out_md.write_text("\n".join(lines), encoding="utf-8")

    print(f"saved -> {OUT_XLSX}  ({OUT_XLSX.stat().st_size} байт)")
    print(f"saved -> {out_md}")
    print(f"  всего артикулов: {len(rows)}")
    for s in ("УНО", "АУРА", "ДИЗАЙН"):
        print(f"  {s}: {sum(1 for r in rows if r['series']==s)}")


if __name__ == "__main__":
    main()
